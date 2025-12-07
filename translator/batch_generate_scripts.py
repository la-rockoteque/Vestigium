#!/usr/bin/env python3
"""
batch_generate_scripts.py

Batch-generates script characters for all phonemes in all scripts.
Similar to batch_generate_lexemes.py but for the Scripts sheet.

Usage:
  python batch_generate_scripts.py                    # Generate for all scripts
  python batch_generate_scripts.py --script=Ashen     # Generate for specific script
  python batch_generate_scripts.py --offline          # Use local XLSX file
  python batch_generate_scripts.py --force            # Overwrite existing cells
"""

import sys
import math
import pandas as pd

from gsheets_client import gsheets, OfflineSheetsClient
from script_generator import llm_script_char

updates = []


def get_client():
  """Choose online or offline client based on flags."""
  offline = "--offline" in sys.argv
  if offline:
    print("Using OFFLINE XLSX client")
    return OfflineSheetsClient("Language.xlsx")
  else:
    print("Using ONLINE Google Sheets client")
    return gsheets


def parse_unicode_range(unicode_str):
  """Extract unicode range from header row 2."""
  if not unicode_str or not isinstance(unicode_str, str):
    return None
  # Remove whitespace and return as-is
  return unicode_str.strip()


def get_target_scripts(df: pd.DataFrame, script_flag: str | None):
  """
  Determine which script columns to process.

  We exclude: Symbol, Category, Place_or_Height, Manner_or_Backness, Voicing, Notes
  Everything else is a script column.
  """
  non_script = {"Symbol", "Category", "Place_or_Height", "Manner_or_Backness", "Voicing", "Notes"}
  all_scripts = [c for c in df.columns if c not in non_script]

  if script_flag is None:
    return all_scripts

  # Case-insensitive matching
  lookup = {c.lower(): c for c in all_scripts}
  wanted = script_flag.lower()
  if wanted not in lookup:
    raise SystemExit(
      f"Script '{script_flag}' not found. "
      f"Available: {', '.join(all_scripts)}"
    )
  return [lookup[wanted]]


def main():
  # Parse optional flags
  script_flag = None
  force = False
  args = sys.argv[1:]
  i = 0
  while i < len(args):
    arg = args[i]
    if arg.startswith("--script="):
      script_flag = arg.split("=", 1)[1].strip()
      i += 1
    elif arg == "--script" and i + 1 < len(args):
      script_flag = args[i + 1].strip()
      i += 2
    elif arg == "--force":
      force = True
      i += 1
    else:
      i += 1

  client = get_client()

  # Load scripts sheet
  df = client.get_df("scripts")

  # The second row (index 0 in data) contains Unicode ranges
  # We need to read this row separately to get the ranges
  # Since get_df returns data starting from row 2 (after headers),
  # we need to handle the unicode ranges manually

  # Read raw records to get the unicode range row
  try:
    ws = client.ws("scripts")
    all_rows = ws.get_all_values()

    # Row 0 is script names, Row 1 is unicode ranges
    if len(all_rows) < 2:
      raise SystemExit("Scripts sheet must have at least 2 header rows (names + unicode ranges)")

    header_row = all_rows[0]
    unicode_row = all_rows[1]

    # Build a mapping of script name -> unicode range
    unicode_ranges = {}
    for i, script_name in enumerate(header_row):
      if i < len(unicode_row):
        unicode_ranges[script_name] = unicode_row[i]

  except AttributeError:
    # Offline mode - read from Excel
    import openpyxl
    wb = openpyxl.load_workbook("Language.xlsx")
    ws = wb["Scripts"]

    header_row = [cell.value for cell in ws[1]]
    unicode_row = [cell.value for cell in ws[2]]

    unicode_ranges = {}
    for i, script_name in enumerate(header_row):
      if i < len(unicode_row):
        unicode_ranges[script_name] = unicode_row[i]

  target_scripts = get_target_scripts(df, script_flag)

  print("Target scripts:", ", ".join(target_scripts))

  # Walk through each phoneme and each target script
  for idx, row in df.iterrows():
    phoneme = str(row.get("Symbol", "")).strip()
    if not phoneme:
      continue

    # Build phoneme info dict
    phoneme_info = {
      'Category': row.get('Category', ''),
      'Place_or_Height': row.get('Place_or_Height', ''),
      'Manner_or_Backness': row.get('Manner_or_Backness', ''),
      'Voicing': row.get('Voicing', ''),
      'Notes': row.get('Notes', '')
    }

    # Sheet row index (2 header rows + df index)
    row_index = idx + 3  # +3 because row 1 is script names, row 2 is unicode ranges

    for script in target_scripts:
      current = row.get(script, "")

      # Treat NaN as empty
      if isinstance(current, float) and math.isnan(current):
        print(f"[INFO] {phoneme!r} [{script}] → Empty, will generate")
        current = ""

      # Skip already-filled cells (unless --force is used)
      if isinstance(current, str) and current.strip():
        if not force:
          print(f"[SKIP] {phoneme!r} [{script}] → Already filled: {current!r} (use --force to overwrite)")
          continue
        else:
          print(f"[INFO] {phoneme!r} [{script}] → Overwriting existing value: {current!r}")

      # Get unicode range for this script
      unicode_range = unicode_ranges.get(script, "")
      if not unicode_range:
        print(f"[WARN] {phoneme!r} [{script}] → No unicode range defined, skipping")
        continue

      # Generate character via LLM
      try:
        char = llm_script_char(phoneme, phoneme_info, script, unicode_range)
        print(f"[INFO] {phoneme!r} [{script}] → Generated: {char!r}")
      except Exception as e:
        print(f"[ERROR] {phoneme!r} [{script}] → Generation failed: {e}")
        continue

      # Compute sheet column (1-based)
      col_index = df.columns.get_loc(script) + 1

      print(f"{phoneme!r} [{script}] → {char!r}  (row {row_index}, col {col_index})")

      # Write back
      updates.append((row_index, col_index, char))

  # Apply batched updates
  if updates:
    print(f"\nApplying {len(updates)} updates...")
    try:
      client.batch_update("scripts", updates)
      print("✓ All updates applied successfully")
    except AttributeError:
      # Fallback if batch_update is not implemented
      print("Batch update not available, applying individually...")
      for r, c, v in updates:
        client.update_cell("scripts", r, c, v)
      print("✓ All updates applied")
  else:
    print("\nNo updates needed - all cells already filled")


if __name__ == "__main__":
  main()
